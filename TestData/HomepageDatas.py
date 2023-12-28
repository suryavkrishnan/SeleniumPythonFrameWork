import openpyxl

class HomePageDatas:
    #METHOD 1
    test_datas = [{"name": "surya", "email": "surya@gmail.com", "passcod": "12345", "gender": "Female", "msg": "hello"},{"name": "pran", "email": "pran@gmail.com", "passcod": "1286434", "gender": "Male", "msg": "Hihello"}]

    # METHOD 2
    @staticmethod
    def getExcelData(testcase_name):
        book = openpyxl.load_workbook("C:\\Users\\Owner\\Desktop\\QAE\\Python\\TestDatas.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        print(cell.value)
        print("-------------")
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        # print(Dict)
        return[Dict]
